import matplotlib.pyplot as plt
import matplotlib as mpl
import numpy as np
import matplotlib.cbook
import warnings

from matplotlib import font_manager

from util.utils import *
from plot.param import *

warnings.filterwarnings("ignore")
import matplotlib.pyplot as plt

font1 = {'family': 'Times New Roman',
         'weight': 'normal',
         'size': 20,
         }
font2 = {'family': 'SimSun',
         'weight': 'normal',
         'size': 20,
         }
font3 = {'family': 'Times New Roman',
         'weight': 'bold',
         # 'style':'italic',
         'size': 20,
         }
figure_no = 1

from xlwt import *
import pandas as pd
from pandas import *
from openpyxl import workbook, load_workbook

# with open("../../result/data_dict/data.xlsx","wb") as f:
#     pass
excel_ = load_workbook("../../result/data_dict/all_result_data.xlsx")


# excel_file = Workbook(encoding = 'utf-8')


# excel = pd.read_excel("../../result/data_dict/data.xlsx")


def draw_line(x: list, y_list: list, label_list: list, color_list: list,
              yticks_labels: list, xticks_label: list,
              marker_list: list, xlimit, ylimit, xlabel: str, ylabel: str,
              figsize: tuple, title: str,
              save_name: str, no: int,
              sci=False, log=False,ncols=1):
    assert len(x) == len(xticks_label)
    assert len(y_list) == len(label_list)
    assert len(y_list) <= len(marker_list)
    assert len(y_list) <= len(color_list)
    x_size = len(x)

    plt.figure(num=no, figsize=figsize, dpi=300)
    plt.rcParams['font.sans-serif'] = ['Times New Roman']
    # plt.rcParams['font.sans-serif'] = ['SimSun','Times New Roman']
    # plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号(无用)

    # 更新公式字体为stix，与times 字体更匹配
    config = {
        "mathtext.fontset": 'stix'
    }
    mpl.rcParams.update(config)

    # 修改刻度线向内 需在plot之前
    plt.rcParams['xtick.direction'] = 'in'
    plt.rcParams['ytick.direction'] = 'in'

    plt.title(title, fontdict=font1)

    # 使用对数曲线
    if log:
        plt.yscale('log')
    # plot 曲线
    for i, y in enumerate(y_list):
        plt.plot(x, y, color=color_list[i], label=label_list[i], marker=marker_list[i])

    plt.yticks(fontproperties='Times New Roman', size=18, ha="right")
    plt.xticks(ticks=np.linspace(x[0], x[-1], x_size), labels=xticks_label, fontproperties='Times New Roman', size=18)

    ax = plt.gca()
    # 设置文字和刻度的间距,防止左下角x和y重叠
    ax.tick_params(axis="x", pad=10)

    # 设置纵坐标科学计数法
    if sci:
        ax.ticklabel_format(style='sci', scilimits=(0, 0), axis='y', useMathText=True)
        print(mpl.rcParams['xtick.labelsize'])
        ax.get_yaxis().get_offset_text().set(va='bottom', ha='left', fontsize="large",
                                             fontproperties="stixgeneral", fontstyle="normal")

    plt.xlabel(xlabel, fontdict=font1)
    plt.ylabel(ylabel, fontdict=font1)
    plt.legend(prop=font3,ncols=ncols)
    plt.grid()
    if xlimit is not None:
        plt.xlim(xlimit[0], xlimit[1])
    if ylimit is not None:
        plt.ylim(ylimit[0], ylimit[1])
    plt.tight_layout()

    f = plt.gcf()
    plt.show()
    if save_name is not None:
        f.savefig("../../result/figure/" + save_name + ".jpg", dpi=300)
        f.savefig("../../result/figure/" + save_name + ".svg", dpi=300, format="svg")


def draw_line_no_cmp(x: list, y: list, xticks_label: list,
                     xlimit, ylimit, xlabel: str, ylabel: str,
                     figsize: tuple, title: str,
                     save_name: str, no: int,
                     sci=False, log=False):
    assert len(x) == len(y)
    x_size = len(x)

    plt.figure(num=no, figsize=figsize, dpi=300)
    plt.rcParams['font.sans-serif'] = ['SimSun', 'Times New Roman']

    # 修改刻度线向内 需在plot之前
    plt.rcParams['xtick.direction'] = 'in'
    plt.rcParams['ytick.direction'] = 'in'

    plt.title(title, fontdict=font1)

    # 使用对数曲线
    if log:
        plt.yscale('log')
    # plot 曲线
    plt.plot(x, y, color="g", marker="o")

    plt.yticks(fontproperties='Times New Roman', size=18, ha="right")
    plt.xticks(ticks=np.linspace(x[0], x[-1], x_size), labels=xticks_label, fontproperties='Times New Roman', size=18)

    ax = plt.gca()
    # 设置文字和刻度的间距,防止左下角x和y重叠
    ax.tick_params(axis="x", pad=10)

    # 设置纵坐标科学计数法
    if sci:
        ax.ticklabel_format(style='sci', scilimits=(0, 0), axis='y', useMathText=True)
        print(mpl.rcParams['xtick.labelsize'])
        ax.get_yaxis().get_offset_text().set(va='bottom', ha='left', fontsize="large",
                                             fontproperties="stixgeneral", fontstyle="normal")

    plt.xlabel(xlabel, fontdict=font1)
    plt.ylabel(ylabel, fontdict=font1)
    # plt.legend(prop=font1)
    plt.grid()
    if xlimit is not None:
        plt.xlim(xlimit[0], xlimit[1])
    if ylimit is not None:
        plt.ylim(ylimit[0], ylimit[1])
    plt.tight_layout()

    f = plt.gcf()
    plt.show()
    if save_name is not None:
        f.savefig("../../result/figure/" + save_name + ".jpg", dpi=300)
        f.savefig("../../result/figure/" + save_name + ".svg", dpi=300, format="svg")


def plot_linear(task, ylimit: list, title_patter: str, sci=False, log=False):
    # task = "HHD_F1"
    cmp_list = cmp_dict[task]
    label_list = cmp_list
    new_label_list = cmp_list[:]
    if "CountMin Sketch" in new_label_list:
        new_label_list.remove("CountMin Sketch")
        new_label_list.append("CM Sketch")

    print(task)
    # table = excel_.create_sheet(task)
    try:
        table = excel_[task]
    except Exception:
        table = excel_.create_sheet(task)
    iii = 2
    # for i, node in enumerate(node_set):
    for i, node in enumerate(node_ll):
        y_list = []

        for algo in cmp_list:
            y_list.append(data_dict[algo][task][node][:data_point_num])
        # title = "Heavy Hitter Detection on " + node.capitalize() + " Node"
        node_name = node_name_map[node]
        title = title_patter + " on " + node_name.capitalize() + " Switches"
        save_name = task + "_" + node_name
        if task == "HHD_ARE":
            ylabel = "ARE"
        elif task == "HHD_F1":
            ylabel = "F1"
        elif task == "CE":
            ylabel = "RE"
        else:
            ylabel = task
        print(node_name)

        table.cell(iii, 1).value = node_name
        iii += 1
        for index, label in enumerate(new_label_list):
            kkk = 1
            print(label + ": ", end=" ")
            table.cell(iii,kkk).value = label + ": "
            kkk += 1
            for yy in y_list[index]:

                print("%.4g" % yy, end=" ")
                table.cell(iii,kkk).value = "%.4g" % yy

                kkk += 1
            iii += 1
            print()
            # print(y_list[index])

        draw_line(x=list(range(data_point_num)), y_list=y_list,
                  label_list=new_label_list, xticks_label=xticks_label, yticks_labels=None,
                  color_list=color_list, marker_list=marker_list,
                  xlimit=[0, data_point_num - 1], ylimit=ylimit,
                  xlabel="Memory Usage (MB)", ylabel=ylabel,
                  figsize=(8, 4.5), title=title,
                  save_name=save_name, no=i, sci=sci, log=log)


def draw_bar_no_cmp(x: list, y: list, xticks_label: list,
                    xlimit, ylimit, xlabel: str, ylabel: str,
                    figsize: tuple, title: str,
                    save_name: str, no: int,
                    sci=False, log=False):
    assert len(x) == len(xticks_label)
    x_size = len(x)

    plt.figure(num=no, figsize=figsize, dpi=300)
    plt.rcParams['font.sans-serif'] = ['SimSun', 'Times New Roman']
    # plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号(无用)
    # 修改刻度线向内 需在plot之前
    plt.rcParams['xtick.direction'] = 'in'
    plt.rcParams['ytick.direction'] = 'in'
    plt.title(title, fontdict=font1)

    # zorder=0 设置图层级别 不遮挡数据
    plt.grid(zorder=0)

    color_list = []
    color_list.extend(['g' for _ in range(8)])
    color_list.extend(['b' for _ in range(8)])
    color_list.extend(['c' for _ in range(4)])
    # 使用对数曲线
    if log:
        plt.yscale('log')
    plt.bar(x, y, width=0.4, zorder=10, color=color_list)
    # plt.bar(x, y, color=color_list[i], label=label_list[i], marker=marker_list[i])

    plt.yticks(fontproperties='Times New Roman', size=18, ha="right")
    plt.xticks(ticks=np.linspace(x[0], x[-1], x_size), labels=xticks_label, fontproperties='Times New Roman',
               size=18, rotation=45)
    ax = plt.gca()

    # 设置文字和刻度的间距,防止左下角x和y重叠
    # ax.tick_params(axis="x", pad=10)
    # 设置纵坐标科学计数法
    if sci:
        ax.ticklabel_format(style='sci', scilimits=(0, 0), axis='y', useMathText=True)
        print(mpl.rcParams['xtick.labelsize'])
        ax.get_yaxis().get_offset_text().set(va='bottom', ha='left', fontsize="large",
                                             fontproperties="stixgeneral", fontstyle="normal")

    plt.xlabel(xlabel, fontdict=font1)
    plt.ylabel(ylabel, fontdict=font1)

    # plt.legend(prop=font1)
    if xlimit is not None:
        plt.xlim(xlimit[0], xlimit[1])
    if ylimit is not None:
        plt.ylim(ylimit[0], ylimit[1])
    plt.tight_layout()

    f = plt.gcf()
    plt.show()
    if save_name is not None:
        f.savefig("../../result/figure/" + save_name + ".jpg", dpi=300)
        f.savefig("../../result/figure/" + save_name + ".svg", dpi=300, format="svg")


# plot_bar():
def draw_error_bar(x: list, y_list: list, e_list: list, label_list: list, color_list: list,
                   yticks_labels: list, xticks_label: list,
                   marker_list: list, xlimit, ylimit, xlabel: str, ylabel: str,
                   figsize: tuple, title: str,
                   save_name: str, no: int,
                   sci=False, log=False):
    assert len(x) == len(xticks_label)
    assert len(y_list) == len(label_list)
    assert len(y_list) <= len(marker_list)
    assert len(y_list) <= len(color_list)
    x_size = len(x)

    plt.figure(num=no, figsize=figsize, dpi=300)
    plt.rcParams['font.sans-serif'] = ['Times New Roman']
    # plt.rcParams['font.sans-serif'] = ['SimSun','Times New Roman']
    # plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号(无用)

    # 修改刻度线向内 需在plot之前
    plt.rcParams['xtick.direction'] = 'in'
    plt.rcParams['ytick.direction'] = 'in'

    plt.title(title, fontdict=font1)

    # 使用对数曲线
    if log:
        plt.yscale('log')
    # plot 曲线
    for i, y in enumerate(y_list):
        print(e_list[i])
        print(y)
        plt.errorbar(x, y, yerr=e_list[i], color=color_list[i], ecolor='b', ms=5, mfc='wheat', mec='salmon', capsize=3,
                     elinewidth=3, label=label_list[i], marker=marker_list[i])
        # plt.plot(x, y, color=color_list[i], label=label_list[i], marker=marker_list[i])

    plt.yticks(fontproperties='Times New Roman', size=18, ha="right")
    plt.xticks(ticks=np.linspace(x[0], x[-1], x_size), labels=xticks_label, fontproperties='Times New Roman', size=18)

    ax = plt.gca()
    # 设置文字和刻度的间距,防止左下角x和y重叠
    ax.tick_params(axis="x", pad=10)

    # 设置纵坐标科学计数法
    if sci:
        ax.ticklabel_format(style='sci', scilimits=(0, 0), axis='y', useMathText=True)
        print(mpl.rcParams['xtick.labelsize'])
        ax.get_yaxis().get_offset_text().set(va='bottom', ha='left', fontsize="large",
                                             fontproperties="stixgeneral", fontstyle="normal")

    plt.xlabel(xlabel, fontdict=font1)
    plt.ylabel(ylabel, fontdict=font1)
    plt.legend(prop=font1)
    plt.grid()
    if xlimit is not None:
        plt.xlim(xlimit[0], xlimit[1])
    if ylimit is not None:
        plt.ylim(ylimit[0], ylimit[1])
    plt.tight_layout()

    f = plt.gcf()
    plt.show()
    if save_name is not None:
        f.savefig("../../result/figure/" + save_name + ".jpg", dpi=300)
        f.savefig("../../result/figure/" + save_name + ".svg", dpi=300, format="svg")


def plot_error_bar(task, ylimit: list, title_patter: str, sci=False, log=False):
    cmp_list = cmp_dict[task]
    label_list = cmp_list
    for i, node in enumerate(node_set):
        y_list = []
        e_list = []

        for algo in cmp_list:
            yt_list = data_dict[algo][task][node][:data_point_num]
            et_list = error_dict[algo][task][node][:data_point_num]
            # et_list.append(error_dict[algo][task][node][:data_point_num])
            # print(error_dict[algo][task][node][:data_point_num])
            # print(et_list)
            et_list = [[yt_list[x] - et_list[x][0] for x in range(data_point_num)],
                       [et_list[x][1] - yt_list[x] for x in range(data_point_num)]]
            # print(et_list)
            y_list.append(yt_list)
            e_list.append(et_list)
        # et_list = [[et_list[x][0] for x in range(data_point_num)], [et_list[x][1] for x in range(data_point_num)]]
        # title = "Heavy Hitter Detection on " + node.capitalize() + " Node"
        title = title_patter + " on " + node.capitalize() + " Node"
        save_name = task + "_" + node + "_error"
        if task == "CE":
            ylabel = "RE"
        else:
            ylabel = task
        print(e_list)
        draw_error_bar(x=list(range(data_point_num)), y_list=y_list, e_list=e_list,
                       label_list=label_list, xticks_label=xticks_label, yticks_labels=None,
                       color_list=color_list, marker_list=marker_list,
                       xlimit=[0, data_point_num - 1], ylimit=ylimit,
                       xlabel="Memory Usage (MB)", ylabel=ylabel,
                       figsize=(8, 4.5), title=title,
                       save_name=save_name, no=i, sci=sci, log=log)


def get_node_list():
    node_list = []
    for i in range(1, 9):
        node_list.append("access" + str(i))
    for i in range(1, 9):
        node_list.append("merge" + str(i))
    for i in range(1, 5):
        node_list.append("core" + str(i))
    return node_list


def plot_statistics_error_bar():
    global data_dict, error_dict
    data_dict = pkl_read("../../result/data_dict/all_data.pkl")
    error_dict = pkl_read("../../result/data_dict/all_error.pkl")

    plot_error_bar("ARE", ylimit=[0, 2], title_patter="Flow Size Measurement")
    # plot_error_bar("HHD_ARE", ylimit=[0, 0.03], title_patter="Heavy Hitter Detection")
    # plot_error_bar("HHD_F1", ylimit=[0.9, 1], title_patter="Heavy Hitter Detection")
    # plot_error_bar("CE", ylimit=[0.00001,0.1], title_patter="Cardinality Estimation",log=True)
    # plot_error_bar("CE", ylimit=[0,0.1], title_patter="Cardinality Estimation",log=False)
    # plot_error_bar("WMRE", ylimit=[0.0001,2], title_patter="Flow Size Distribution Estimation",log=True)


def plot_statistics():
    # global data_dict
    # for algo in algo_set:
    #     data_dict[algo] = pkl_read(data_dict_dir + algo + ".pkl")
    # data_dict.update(pkl_read("../../result/data_dict/output.pkl"))
    # print(data_dict.keys())
    # data_dict["TM-SALFM"] = data_dict["ServerCount Sketch"]
    # data_dict.pop("ServerCount Sketch")
    # print(data_dict.keys())
    #
    #
    #
    # pkl_write("../../result/data_dict/all_data.pkl", data_dict)
    global data_dict
    data_dict = pkl_read("../../result/data_dict/all_data.pkl")

    # print(data_dict)
    # for d,v in data_dict.items():
    #     print(d)
    #     for dd,vv in v.items():
    #         print(dd)
    #         for ddd,vvv in vv.items():
    #             print(ddd)
    #             print(vvv)

    # plot_linear("ARE", ylimit=[0, 2], title_patter="Flow Size Measurement")
    # plot_linear("HHD_ARE", ylimit=[0, 0.03], title_patter="Heavy Hitter Detection")
    # plot_linear("HHD_F1", ylimit=[0.90, 1], title_patter="Heavy Hitter Detection")
    plot_linear("CE", ylimit=[0, 0.03], title_patter="Cardinality Estimation", log=False)
    # plot_linear("WMRE", ylimit=[0.0005, 2], title_patter="Flow Size Distribution Estimation", log=True)
    # plot_linear("CE", ylimit=[0.00001,0.1], title_patter="Cardinality Estimation",log=True)
    # excel_.save("../../result/data_dict/data1.xlsx")
    # excel_.save("../../result/data_dict/data1.xlsx")


def plot_flow_and_pkt_count():
    flow_count_dict = pkl_read("../../result/data_dict/flow_count_per_node_dict.pkl")
    pkt_count_dict = pkl_read("../../result/data_dict/pkt_count_per_node_dict")

    flow_count_list = []
    pkt_count_list = []

    node_list = get_node_list()
    # node_name = node_name_map[node]
    for node in node_list:
        flow_count_list.append(flow_count_dict[node])
        pkt_count_list.append(pkt_count_dict[node])

    # change xticks_label
    xticks_label = []
    for node in node_list:
        if node[:-1] == "access":
            xticks_label.append("Edge" + node[-1])
        elif node[:-1] == "merge":
            xticks_label.append("Agg" + node[-1])
        else:
            xticks_label.append(node)

    print(xticks_label)
    print(len(xticks_label))
    print(flow_count_dict)
    print(pkt_count_dict)
    draw_bar_no_cmp(x=list(range(20)), y=flow_count_list,
                    xlimit=[-1, 20], ylimit=[0, 90000], xticks_label=xticks_label,
                    xlabel="Node", ylabel="Number of Flows",
                    figsize=(9, 4.5), title="Number of Flows on Each Node",
                    save_name="Flows_Number", no=1, sci=True, log=False)

    draw_bar_no_cmp(x=list(range(20)), y=pkt_count_list,
                    xlimit=[-1, 20], ylimit=[0, 500000], xticks_label=xticks_label,
                    xlabel="Node", ylabel="Number of Packets",
                    figsize=(9, 4.5), title="Number of Packets on Each Node",
                    save_name="Packets_Number", no=1, sci=True, log=False)


def plot_CDF():
    cdf_list = pkl_read("../../result/data_dict/cdf_data.pkl")
    # plot_linear("Percentage", ylimit=[0.0001, 2], title_patter="Cumulative Distribution Function", log=True)

    draw_line_no_cmp(x=list(range(0, 101, 5)), y=[int(cdf_list[i] * 100) for i in range(0, 101, 5)],
                     xticks_label=[i for i in range(0, 101, 5)],
                     xlimit=[0, 100], ylimit=[0, 100],
                     xlabel="Percentage of Top Flow(%)", ylabel="Percentage of Packets(%)",
                     figsize=(8, 4.5), title="Cumulative Distribution Function of Flow Size",
                     save_name="CDF", no=1, sci=False, log=False)


def get_all_data():
    global data_dict
    data_dict["TM-SALFI"] = pkl_read("../../result/data_dict/TM-SALFI.pkl")
    data_dict.update(pkl_read("../../result/data_dict/output_10.pkl"))
    # data_dict.update(pkl_read("../../result/data_dict/edge.pkl"))

    print(pkl_read("../../result/data_dict/output_10.pkl"))
    edge_dict = pkl_read("../../result/data_dict/edge.pkl")
    print(edge_dict)
    for al,d in edge_dict.items():
        for task,dd in d.items():
            data_dict[al][task]["access"] = edge_dict[al][task]["access"]
    # print(pkl_read("../../result/data_dict/edge.pkl"))
    print(data_dict)
    print(data_dict.keys())
    data_dict["TM-SALFM"] = data_dict["ServerCount Sketch"]
    data_dict.pop("ServerCount Sketch")
    global error_dict
    error_dict = pkl_read("../../result/data_dict/error_1.pkl")
    error_dict.update(pkl_read("../../result/data_dict/confidence_interval.pkl"))
    error_dict["TM-SALFM"] = error_dict["ServerCount Sketch"]
    error_dict.pop("ServerCount Sketch")
    print(error_dict.keys())
    pkl_write("../../result/data_dict/all_data.pkl", data_dict)
    pkl_write("../../result/data_dict/all_error.pkl", error_dict)


def plot_kw_cmp():
    data = pkl_read("../../result/data_dict/kw_cmp.pkl")

    k_list = ["2", "4", "8"]
    label_list = ["k = " + k for k in k_list]
    y_list = [data[k] for k in k_list]
    xticks_label = [1, 0.5, 0.2, 0.1, 0.05, 0.02, 0.01]
    xticks_label = [str(x) for x in xticks_label]
    # xlimit = [0]
    # r"N/(k$\times$m)"
    draw_line(x=list(range(7)), y_list=y_list,
              label_list=label_list, xticks_label=xticks_label, yticks_labels=None,
              color_list=color_list, marker_list=marker_list,
              xlimit=[0, 6], ylimit=[0.05293, 0.05308],
              xlabel=r"N/(k$\times$m)", ylabel="ARE",
              figsize=(8, 4.5), title="Flow Size Measurement On Edge Switches",
              save_name="Nkm_cmp", no=0, sci=True, log=False)

def plot_diff_t():
    obj1 = [
        [0.293344, 0.13353, 0.071483, 0.043326, 0.027586, 0.019611, 0.014328, 0.011046, 0.008357],
        [0.216156, 0.095686, 0.052105, 0.031021, 0.019396, 0.014798, 0.010632, 0.007998, 0.006094],
        [0.14748, 0.063228, 0.034182, 0.020809, 0.012727, 0.009925, 0.007076, 0.005412, 0.004238],
        [0.093615, 0.039474, 0.021252, 0.013086, 0.00789, 0.006142, 0.004621, 0.003352, 0.002718],
        [0.054404, 0.022221, 0.01214, 0.00771, 0.004837, 0.003939, 0.002885, 0.002047, 0.001868]
    ]

    label_list = ["t = "+str(i) + "%" for i in range(10,60,10)]

    # ncols 是 legend图例显示的列数 可以让我们一行显示多个图例
    draw_line(x=list(range(data_point_num)), y_list=obj1,
              label_list=label_list, xticks_label=xticks_label, yticks_labels=None,
              color_list=color_list, marker_list=marker_list,
              xlimit=[0, data_point_num - 1], ylimit=[0,0.3],
              xlabel="Memory Usage (MB)", ylabel="ARE",
              figsize=(8, 4.5), title="Flow Size Measurement on Edge Switches",
              save_name="t_cmp", no=0, sci=False, log=False,ncols=2)

if __name__ == '__main__':
    # get_all_data()

    # global data_dict
    # for algo in algo_set:
    #     data_dict[algo] = pkl_read(data_dict_dir + algo + ".pkl")
    # data_dict.update(pkl_read("../../result/data_dict/output.pkl"))

    # for k,v in data_dict.items():
    #     print(k)
    #     print(v)
    # print(data_dict[algo])
    # print(data_dict)
    # plot_linear("ARE", ylimit=[0, 2], title_patter="Flow Size Measurement")
    # plot_linear("HHD_ARE", ylimit=[0, 0.03], title_patter="Heavy Hitter Detection")
    # plot_linear("HHD_F1", ylimit=[0.9, 1], title_patter="Heavy Hitter Detection")
    # plot_linear("CE", ylimit=[0.00001,0.1], title_patter="Cardinality Estimation",log=True)
    # plot_linear("CE", ylimit=[0,0.1], title_patter="Cardinality Estimation",log=False)
    # plot_linear("WMRE", ylimit=[0.0001,2], title_patter="Flow Size Distribution Estimation",log=True)

    # plot_flow_and_pkt_count()
    # plot_statistics()
    # plot_statistics_error_bar()
    # plot_kw_cmp()
    plot_diff_t()
